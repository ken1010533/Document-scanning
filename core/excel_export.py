"""Excel 匯出工具。

目前用來匯出重複檔案群組；UI 會決定輸出位置與是否自動開啟。
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import subprocess
import platform


def 輸出Excel(重複資料, 輸出檔="重複檔案.xlsx", 自動開啟=True):
    """把重複檔案群組寫成 Excel，並可選擇輸出後自動開啟。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "重複檔案"

    if not 重複資料:
        print("沒有重複資料")
        return

    ws.append(["群組ID", "檔案路徑", "檔案大小(MB)", "所在資料夾"])

    顏色 = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    群組ID = 1

    for 群組 in 重複資料:
        for 檔案 in 群組:
            try:
                路徑 = os.path.normpath(os.path.abspath(檔案))
                大小MB = round(os.path.getsize(路徑) / 1024 / 1024, 2)
                資料夾 = os.path.dirname(路徑)

                ws.append([群組ID, 路徑, 大小MB, 資料夾])

                row = ws.max_row

                cell = ws.cell(row=row, column=4)
                cell.hyperlink = 資料夾
                cell.value = 資料夾
                cell.style = "Hyperlink"

                if 群組ID % 2 == 0:
                    for c in ws[row]:
                        c.fill = 顏色

            except:
                continue

        群組ID += 1

    # 自動欄寬
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[letter].width = max_len + 2

    ws.freeze_panes = "A2"

    wb.save(輸出檔)
    print(f"Excel已輸出：{輸出檔}")

    if 自動開啟:
        開啟檔案(輸出檔)


def 開啟檔案(路徑):
    """依作業系統使用預設程式開啟輸出檔。"""
    系統 = platform.system()

    if 系統 == "Windows":
        os.startfile(路徑)
    elif 系統 == "Darwin":
        subprocess.call(["open", 路徑])
    else:
        subprocess.call(["xdg-open", 路徑])
